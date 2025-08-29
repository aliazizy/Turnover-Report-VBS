import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------- CONFIG ----------
JSON_FILE = "turnover.json"
OUTPUT_XLSX = "turnover-output.xlsx"

# ---------- LOAD JSON ----------
with open(JSON_FILE, encoding="utf-8") as f:
    data = json.load(f)

caption = data["caption"]  # "Turnover Report"
date_time_user = data["dateTimeUser"]
legend = f"{data['legend'][0]['label']}: {data['legend'][0]['value']}"
columns = [c["caption"] for c in data["columns"]]  # Display headers
rows = data["rows"]                                # List of row dicts

# ---------- BUILD WORKBOOK ----------
wb = Workbook()
ws = wb.active
ws.title = "Report"

# ---------- HEADER AREA ----------
# A1: DateTimeUser
cell = ws["A1"]
cell.value = date_time_user
cell.fill = PatternFill("solid", fgColor="FFFF00")   # yellow fill
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal="left")

# A2: Caption
cell = ws["A2"]
cell.value = caption
cell.font = Font(size=14, bold=True)
cell.alignment = Alignment(horizontal="left")

# A3: Legend
cell = ws["A3"]
cell.value = legend
cell.font = Font(italic=True)
cell.alignment = Alignment(horizontal="left")

# ---------- TABLE ----------
header_row = 5
# Write headers
for col_idx, header in enumerate(columns, start=1):
    c = ws.cell(row=header_row, column=col_idx, value=header)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4F81BD")  # blue fill
    c.alignment = Alignment(horizontal="center")

# Write data rows
for r, row in enumerate(rows, start=header_row+1):
    for c, col_def in enumerate(data["columns"], start=1):
        val = row.get(col_def["name"], "")
        ws.cell(row=r, column=c).value = val

# ---------- FORMAT COLUMNS ----------
end_row = header_row + len(rows)
end_col = len(columns)

# Adjust column widths based on header + content
for idx, col_def in enumerate(data["columns"], start=1):
    col_caption = col_def["caption"]
    col_key = col_def["name"]

    max_len = max(
        len(str(col_caption)),
        max((len(str(r.get(col_key, ""))) for r in rows), default=0)
    )
    ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

# EUR column formatting
eur_col = columns.index("Turnover\r\nEUR") + 1
for r in range(header_row+1, end_row+1):
    ws.cell(row=r, column=eur_col).number_format = "#,##0.00"

# % column formatting
pct_col = columns.index("%") + 1
for r in range(header_row+1, end_row+1):
    ws.cell(row=r, column=pct_col).number_format = "0.0%"

# ---------- CREATE EXCEL TABLE ----------
table_range = f"A{header_row}:{get_column_letter(end_col)}{end_row}"
table = Table(displayName="TurnoverTable", ref=table_range)

# Blue header + banded rows
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
table.tableStyleInfo = style
ws.add_table(table)

# ---------- SAVE ----------
wb.save(OUTPUT_XLSX)
print(f"✅ Excel file written: {OUTPUT_XLSX}")
