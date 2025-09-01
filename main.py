# main.py
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware

import io, json, logging, time
from typing import List, Dict, Any
from openpyxl.drawing.image import Image as XLImage

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os
from pathlib import Path

LOG = logging.getLogger("turnover-api")

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_LOGO_PATH = BASE_DIR / "logo.png"              # bundled file
ENV_LOGO_PATH = os.getenv("LOGO_PATH")                 # e.g., /home/site/wwwroot/logo.png
ENV_LOGO_URL  = os.getenv("LOGO_URL")                  # e.g., https://.../logo.png (SAS/Blob/CDN)

# If a URL is provided, fetch once into temp file
_cached_logo_file = None
if ENV_LOGO_URL:
    try:
        import requests, tempfile
        resp = requests.get(ENV_LOGO_URL, timeout=10)
        resp.raise_for_status()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        tmp.write(resp.content); tmp.flush(); tmp.close()
        _cached_logo_file = Path(tmp.name)
        LOG.info(f"Downloaded logo from URL to {_cached_logo_file}")
    except Exception as e:
        LOG.warning(f"Failed to download logo from LOGO_URL: {e}")

def resolve_logo_path() -> Path | None:
    if _cached_logo_file and _cached_logo_file.exists():
        return _cached_logo_file
    if ENV_LOGO_PATH and Path(ENV_LOGO_PATH).exists():
        return Path(ENV_LOGO_PATH)
    if DEFAULT_LOGO_PATH.exists():
        return DEFAULT_LOGO_PATH
    return None

# ---------------- Logging ----------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(name)s - %(message)s"
)
logger = logging.getLogger("turnover-api")

# ---------------- App ----------------
app = FastAPI(
    title="Turnover Report Generator",
    version="1.0.0",
    description=(
        "Upload a turnover JSON and receive a styled Excel file that matches the template screenshot:\n"
        "- A1:D1 merged yellow banner\n"
        "- Blue title in A2, legend in A3\n"
        "- Row 5 light-blue header with AutoFilter\n"
        "- Frozen panes so headers stay visible\n"
        "- Proper currency & percent formats\n"
        "- Totals row with thick border\n"
        "- Logo from logo.png near H2"
    ),
    openapi_url="/openapi.json",
    docs_url="/docs",
    redoc_url="/redoc",
)

# Request logging middleware
@app.middleware("http")
async def log_requests(request: Request, call_next):
    rid = f"{time.time_ns()}"
    logger.info(f"rid={rid} start path={request.url.path} method={request.method}")
    start = time.time()
    try:
        response = await call_next(request)
    except Exception as ex:
        logger.exception(f"rid={rid} exception: {ex}")
        raise
    finally:
        took_ms = (time.time() - start) * 1000.0
    logger.info(f"rid={rid} done path={request.url.path} status={response.status_code} took_ms={took_ms:.2f}")
    return response

# CORS (open for dev; restrict in prod)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------- Excel builder ----------------
def build_workbook(payload: Dict[str, Any]) -> io.BytesIO:
    """
    Build an .xlsx that visually matches the screenshot:
      - merged yellow banner A1:D1
      - title (blue, larger, bold) in A2; legend in A3
      - header row 5 with light blue fill + AutoFilter (no table)
      - freeze panes at A6
      - currency & percent formats
      - totals row with thick top border
      - logo.png inserted around H2
    """
    # Validate minimal shape
    for key in ("caption", "dateTimeUser", "legend", "columns", "rows"):
        if key not in payload:
            raise ValueError(f"Missing required key: '{key}'")

    caption: str = payload["caption"]
    date_time_user: str = payload["dateTimeUser"]
    legend_items = payload["legend"]

    legend_text = ""
    if isinstance(legend_items, list) and legend_items:
        legend_text = " | ".join([f"{item.get('label','')}: {item.get('value','')}" for item in legend_items])

    # Columns / rows
    col_defs: List[Dict[str, str]] = payload["columns"]
    rows: List[Dict[str, Any]] = payload["rows"]

    if not col_defs:
        raise ValueError("No columns provided.")

    captions: List[str] = [c.get("caption", "") for c in col_defs]   # display headers
    keys: List[str] = [c.get("name", "") for c in col_defs]          # JSON keys

    # Colors (picked to match the screenshot closely)
    YELLOW_BANNER = "FFFF00"   # row 1
    TITLE_BLUE = "1F4E78"      # title font color
    HEADER_FILL = "DDEBF7"     # light blue header band
    HEADER_FONT = "000000"     # black header text

    # Borders
    thin = Side(style="thin", color="9E9E9E")
    thick = Side(style="thick", color="000000")

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Column count
    end_col = len(captions)
    last_col_letter = get_column_letter(end_col)

    # ----- Row 1: merged yellow banner -----
    ws.merge_cells(f"A1:{last_col_letter}1")
    a1 = ws["A1"]
    a1.value = date_time_user
    a1.fill = PatternFill("solid", fgColor=YELLOW_BANNER)
    a1.font = Font(bold=True)
    a1.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # ----- Row 2: Title -----
    ws.merge_cells(f"A2:{last_col_letter}2")
    a2 = ws["A2"]
    a2.value = caption
    a2.font = Font(size=16, bold=True, color=TITLE_BLUE)
    a2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # ----- Row 3: Legend -----
    ws.merge_cells(f"A3:{last_col_letter}3")
    a3 = ws["A3"]
    a3.value = legend_text
    a3.font = Font(italic=True)
    a3.alignment = Alignment(horizontal="left", vertical="center")

    # Row 4 is a spacer (as per screenshot look)
    ws.row_dimensions[4].height = 6

    # ----- Row 5: Header band (light blue) -----
    header_row = 5
    for col_idx, header in enumerate(captions, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=header)
        c.font = Font(bold=True, color=HEADER_FONT)
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.row_dimensions[header_row].height = 20

    # ----- Data rows -----
    # Identify special columns
    percent_col_idx = None
    turnover_col_idx = None
    for idx, cap in enumerate(captions, start=1):
        if cap.strip() == "%":
            percent_col_idx = idx
        norm = cap.replace("\r", "").replace("\n", "").strip().lower()
        if norm.startswith("turnover"):
            turnover_col_idx = idx

    start_row = header_row + 1
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, key in enumerate(keys, start=1):
            val = row.get(key, "")
            if percent_col_idx is not None and c_idx == percent_col_idx:
                try:
                    if val not in (None, ""):
                        val = float(val) / 100.0  # 6.4 -> 0.064
                except Exception:
                    pass
            if turnover_col_idx is not None and c_idx == turnover_col_idx:
                try:
                    if val not in (None, ""):
                        val = float(val)
                except Exception:
                    pass
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = Border(left=thin, right=thin)

    data_rows = len(rows)
    end_row = header_row + data_rows

    # Column widths (auto-ish)
    for idx, col_def in enumerate(col_defs, start=1):
        cap = col_def["caption"]; key = col_def["name"]
        max_len_in_data = 0
        for row in rows:
            max_len_in_data = max(max_len_in_data, len(str(row.get(key, ""))))
        ws.column_dimensions[get_column_letter(idx)].width = max(len(str(cap)), max_len_in_data) + 5
    
    

    # Number formats
    if turnover_col_idx:
        for r in range(start_row, end_row + 1):
            ws.cell(row=r, column=turnover_col_idx).number_format = "#,##0.00"
    if percent_col_idx:
        for r in range(start_row, end_row + 1):
            ws.cell(row=r, column=percent_col_idx).number_format = "0.0%"

    # ----- AutoFilter (filters arrows like screenshot) -----
    if data_rows >= 0:
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{max(end_row, header_row)}"

    # ----- Freeze header row -----
    ws.freeze_panes = ws["A6"]  # row 5 fixed

    # ----- Totals row (bold + thick top border) -----
    total_row = end_row + 1 if data_rows >= 1 else header_row + 1
    if turnover_col_idx:
        # Label in column before turnover if available, else in A
        label_col = turnover_col_idx - 1 if turnover_col_idx > 1 else 1
        ws.cell(row=total_row, column=label_col).value = "Total"
        ws.cell(row=total_row, column=label_col).font = Font(bold=True)

        # Sum formulas
        first_data = start_row
        last_data = end_row if data_rows >= 1 else start_row  # safe
        sum_cell = ws.cell(row=total_row, column=turnover_col_idx)
        sum_cell.value = f"=SUM({get_column_letter(turnover_col_idx)}{first_data}:{get_column_letter(turnover_col_idx)}{last_data})"
        sum_cell.number_format = "#,##0.00"
        sum_cell.font = Font(bold=True)

        if percent_col_idx:
            pct_sum_cell = ws.cell(row=total_row, column=percent_col_idx)
            pct_sum_cell.value = f"=SUM({get_column_letter(percent_col_idx)}{first_data}:{get_column_letter(percent_col_idx)}{last_data})"
            pct_sum_cell.number_format = "0.0%"
            pct_sum_cell.font = Font(bold=True)

        # Thick top border across full width
        for c_idx in range(1, end_col + 1):
            cell = ws.cell(row=total_row, column=c_idx)
            cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)

    # ----- Logo (logo.png) near H2 (if present) -----
    try:
        logo_path = resolve_logo_path()
        if logo_path:
            img = XLImage(str(logo_path))
            # exact size ~ 2.9" x 0.42" at 96 DPI
            img.width = 278  # px
            img.height = 40  # px
            ws.add_image(img, "H2")
            logger.info(f"Logo added from {logo_path}")
        else:
            logger.info("Logo not found (no DEFAULT/ENV/URL). Skipping.")
    except Exception as e:
        logger.warning(f"Logo not added: {e}")
    # ----- Save & self-validate -----
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    try:
        _ = load_workbook(buf)  # validation
        buf.seek(0)
    except Exception as e:
        raise ValueError(f"Generated workbook failed validation: {e}")

    return buf


# ---------------- Routes ----------------
@app.post(
    "/report",
    summary="Upload turnover JSON and receive styled Excel file",
    tags=["Report"],
    responses={
       200: {"content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}}},
       400: {"description": "Bad Request"},
       500: {"description": "Server Error"},
    },
    openapi_extra={
        "requestBody": {
            "content": {
                "multipart/form-data": {
                    "schema": {
                        "type": "object",
                        "properties": {"file": {"type": "string", "format": "binary"}},
                        "required": ["file"]
                    }
                }
            }
        }
    },
)
async def create_report(file: UploadFile = File(..., description="Upload a `turnover.json` file")):
    """
    Upload a turnover JSON and receive an Excel file matching the template screenshot:
    - merged banner, header band with filters, frozen headers, logo, totals row
    """
    if file.content_type not in ("application/json", "text/json", "application/octet-stream"):
        raise HTTPException(status_code=400, detail="Please upload a JSON file.")
    logger.info(f"/report upload: filename={file.filename} content_type={file.content_type}")

    try:
        raw = await file.read()
        payload = json.loads(raw.decode("utf-8"))
        xlsx_bytes = build_workbook(payload)
    except ValueError as ve:
        logger.warning(f"/report bad request: {ve}")
        raise HTTPException(status_code=400, detail=str(ve))
    except json.JSONDecodeError:
        logger.warning("/report invalid JSON")
        raise HTTPException(status_code=400, detail="Invalid JSON.")
    except Exception as ex:
        logger.exception(f"/report failed: {ex}")
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {ex}")

    headers = {"Content-Disposition": 'attachment; filename="turnover-report.xlsx"'}
    return StreamingResponse(
        xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )

@app.get("/", include_in_schema=False)
def root():
    return JSONResponse({"status": "ok", "upload_to": "/report", "docs": "/docs", "redoc": "/redoc"})
